# -*- coding: utf-8 -*-
"""
Utilitaires pour l'importation de données depuis Excel
Adapté à la structure du fichier "Liste Tireurs Tir des Amis.xlsx"
"""

from pathlib import Path
from typing import Dict, List, Tuple
from openpyxl import load_workbook
from datetime import datetime, date
from database.models import Database, Shooter, Result, Competition
from config.settings import EXCEL_CONFIG

def import_shooters_from_excel(file_path: str, db: Database) -> Dict:
    """
    Importe les tireurs depuis un fichier Excel
    Structure attendue : Nom | Prénom | Société | Année | Arme | Âge | Score | Série | N.10 | Prix
    
    Args:
        file_path: Chemin vers le fichier Excel
        db: Instance de la base de données
    
    Returns:
        Dict avec le statut de l'import et les détails
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet_name = 'Tireurs'  # Première feuille
        
        if sheet_name not in wb.sheetnames:
            # Essayer la première feuille si 'Tireurs' n'existe pas
            sheet_name = wb.sheetnames[0]
        
        ws = wb[sheet_name]
        imported_count = 0
        errors = []
        results_data = []
        
        # Parcourir les lignes (en ignorant l'en-tête - min_row=2)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # Vérifier que la ligne n'est pas vide
                if not any(row):
                    continue
                
                # Extraire les colonnes selon la structure observée
                # Colonnes : Rang(0) | Nom(1) | Prénom(2) | Société(3) | Année(4) | Arme(5) | Âge(6) | Score(7) | Série(8) | N.10(9) | Prix(10)
                
                nom = row[1]  # Colonne B
                prenom = row[2]  # Colonne C
                societe = row[3]  # Colonne D
                annee_naissance = row[4]  # Colonne E
                arme = row[5]  # Colonne F
                age = row[6]  # Colonne G
                score = row[7]  # Colonne H
                serie = row[8]  # Colonne I
                n10 = row[9]  # Colonne J
                
                # Valider les données essentielles
                if not nom or not prenom:
                    errors.append(f"Ligne {row_idx}: Nom ou prénom manquant")
                    continue
                
                # Convertir l'année en date de naissance
                if isinstance(annee_naissance, (int, float)):
                    annee = int(annee_naissance)
                    # Créer une date approximative : 1er janvier de l'année
                    date_naissance = date(annee, 1, 1)
                else:
                    errors.append(f"Ligne {row_idx}: Année invalide")
                    continue
                
                # Créer l'objet Shooter
                shooter = Shooter(
                    nom=str(nom).strip(),
                    prenom=str(prenom).strip(),
                    date_naissance=date_naissance,
                    sexe='',  # À déterminer depuis une autre feuille si disponible
                    arme=str(arme).strip() if arme else '',
                    calibre='',  # À récupérer depuis les listes
                    cartouches=0,
                    societe=str(societe).strip() if societe else ''\n                )\n                
                # Ajouter le tireur\n                added_shooter = db.add_shooter(shooter)\n                imported_count += 1\n                \n                # Stocker les résultats pour création ultérieure\n                if score or serie or n10:\n                    results_data.append({\n                        'shooter_id': added_shooter.id,\n                        'points': float(score) if score else 0.0,\n                        'feu_de_serie': int(serie) if serie else 0,\n                        'nombre_10': int(n10) if n10 else 0\n                    })\n            \n            except Exception as e:\n                errors.append(f\"Ligne {row_idx}: {str(e)}\")\n        \n        return {\n            'success': True,\n            'imported': imported_count,\n            'errors': errors,\n            'results': results_data\n        }\n    \n    except Exception as e:\n        return {\n            'success': False,\n            'error': str(e),\n            'imported': 0\n        }\n\n\ndef import_reference_data(file_path: str) -> Dict:\n    \"\"\"\n    Importe les données de référence (catégories, armes, calibres, sociétés, prix)\n    \n    Args:\n        file_path: Chemin vers le fichier Excel\n    \n    Returns:\n        Dict avec les données importées\n    \"\"\"\n    try:\n        wb = load_workbook(file_path, data_only=True)\n        data = {}\n        \n        # Importer les listes si les feuilles existent\n        sheet_names = wb.sheetnames\n        \n        if 'Listes' in sheet_names:\n            data['listes'] = _import_reference_sheet(wb['Listes'])\n        \n        if 'Armes' in sheet_names:\n            data['armes'] = _import_simple_list(wb['Armes'])\n        \n        if 'Calibres' in sheet_names:\n            data['calibres'] = _import_simple_list(wb['Calibres'])\n        \n        if 'Sociétés' in sheet_names:\n            data['societes'] = _import_simple_list(wb['Sociétés'])\n        \n        if 'Prix' in sheet_names:\n            data['prix'] = _import_price_list(wb['Prix'])\n        \n        return {'success': True, 'data': data}\n    \n    except Exception as e:\n        return {'success': False, 'error': str(e)}\n\n\ndef _import_reference_sheet(sheet) -> Dict:\n    \"\"\"\n    Importe une feuille de références (catégories, etc.)\n    \"\"\"\n    data = {}\n    \n    for row in sheet.iter_rows(min_row=2, values_only=True):\n        if row[0]:  # Si la première colonne n'est pas vide\n            key = str(row[0]).strip()\n            value = str(row[1]).strip() if len(row) > 1 and row[1] else ''\n            data[key] = value\n    \n    return data\n\n\ndef _import_simple_list(sheet) -> List[str]:\n    \"\"\"\n    Importe une simple liste depuis une feuille Excel\n    \"\"\"\n    items = []\n    for row in sheet.iter_rows(min_row=2, values_only=True):\n        if row[0]:\n            items.append(str(row[0]).strip())\n    return items\n\n\ndef _import_price_list(sheet) -> Dict:\n    \"\"\"\n    Importe la liste des prix (Rang -> Montant)\n    \"\"\"\n    prices = {}\n    for row in sheet.iter_rows(min_row=2, values_only=True):\n        if row[0] and row[1]:\n            try:\n                rang = int(row[0])\n                montant = float(row[1])\n                prices[rang] = montant\n            except (ValueError, TypeError):\n                continue\n    return prices\n